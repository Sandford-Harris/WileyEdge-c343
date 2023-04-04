import openpyxl

class Course:
    def __init__(self, cid="", desc=""):
        self.cid = cid
        self.desc = desc

    def getCourseId(self):
        return self.cid

    def setCourseId(self, cid):
        self.cid = cid

    def getDescription(self):
        return self.desc

    def setDescription(self, desc):
        self.desc = desc

    def findCourseInList(self, ws) :  # returns row in the sheet that the course is in
        if ws.title == "CourseDetails":
            column = 1
        else:
            column = 2
        values = [i for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=column).value == self.cid]
        if len(values) > 0 :
            return values[0]
        else :
            return False

    def checkExistingCourses(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["CourseDetails"]
        if self.findCourseInList(ws):
            return True
        else :
            return False

    def saveToExcel(self):
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["CourseDetails"]
        ws.append((self.cid, self.desc))

        wb.save("ManagementSystem.xlsx")

    def deleteCourse(self):  # deletes a course that has no trainees or trainers assigned
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["MappingCourseTrainee"]
        if self.findCourseInList(ws):
            return False
        ws = wb["MappingCourseTrainer"]
        if self.findCourseInList(ws):
            return False

        ws = wb["CourseDetails"]
        row = self.findCourseInList(ws)
        if row:
            ws.delete_rows(row, 1)
            wb.save("ManagementSystem.xlsx")
            return True

        return False

    def updateCourse(self, desc):
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["CourseDetails"]
        row = self.findCourseInList(ws)
        if row:
            if desc != "" :
                ws.cell(row=row, column=2).value = desc
            wb.save("ManagementSystem.xlsx")
            return True
        else :
            return False

    def getTrainerId(self):
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["MappingCourseTrainer"]
        row = self.findCourseInList(ws)
        if row:
            return ws.cell(row=row, column=1).value
        else :
            return False

    def getTraineeIdList(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["MappingCourseTrainee"]
        values = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=2).value == self.cid]
        return values

    def removeTrainer(self):
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["MappingCourseTrainer"]
        row = self.findCourseInList(ws)
        if row:
            tid = self.getTrainerId()
            ws.delete_rows(row, 1)
            wb.save("ManagementSystem.xlsx")
            return tid
        else :
            return False

    def setTrainer(self, tid) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["MappingCourseTrainer"]
        row = self.findCourseInList(ws)
        if row :
            ws.cell(row=row, column=1).value = tid
            return True
        else:
            return False
