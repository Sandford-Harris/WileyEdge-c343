import openpyxl
from Course import Course

class Trainer:
    def __init__(self, tid="", name="", course="", email="", phone=""):
        self.trainerId = tid
        self.name = name
        self.course = course
        self.email = email
        self.phoneNumber = phone

    def getId(self):
        return self.trainerId

    def setId(self, tid):
        self.trainerId = tid

    def getName(self):
        return self.name

    def setName(self, name):
        self.name = name

    def getCourse(self) :
        return self.course

    def setCourse(self, course) :
        self.course = course

    def getPhoneNumber(self):
        return self.phoneNumber

    def setPhoneNumber(self, phone):
        self.phoneNumber = phone

    def getEmail(self):
        return self.email

    def setEmail(self, email):
        self.email = email

    def findTrainerInList(self, ws) :  # returns row in the sheet that the employee is in
        values = [i for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=1).value == self.trainerId]
        if len(values) > 0:
            return values[0]
        else :
            return False

    def checkExistingTrainer(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["TrainerDetails"]
        if self.findTrainerInList(ws):
            return True
        else :
            return False

    def saveToExcel(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["TrainerDetails"]

        newRow = (self.trainerId, self.name, self.email, self.phoneNumber)
        ws.append(newRow)

        ws = wb["MappingCourseTrainer"]
        ws.append((self.trainerId, self.course))

        wb.save("ManagementSystem.xlsx")

    def deleteTrainer(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        deleted = False

        ws = wb["TrainerDetails"]  # remove trainee from list of trainees
        row2del = self.findTrainerInList(ws)
        if row2del:
            ws.delete_rows(row2del, 1)
            deleted = True

        ws = wb["MappingCourseTrainer"]  # remove trainee from list of course mappings
        row2del = self.findTrainerInList(ws)
        if row2del:
            ws.delete_rows(row2del, 1)
            deleted = True

        wb.save("ManagementSystem.xlsx")
        return deleted

    def updateExistingTrainer(self, course, name, email, phone) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["TrainerDetails"]
        row = self.findTrainerInList(ws)
        if row:
            if name != "":
                ws.cell(row=row, column=2).value = name
            if email != "":
                ws.cell(row=row, column=3).value = email
            if phone != "":
                ws.cell(row=row, column=4).value = phone

            if course != "":
                c = Course(cid=course)
                if c.checkExistingCourses():
                    ws = wb["MappingCourseTrainer"]
                    row = self.findTrainerInList(ws)
                    if row:
                        ws.cell(row=row, column=2).value = course
                    else:
                        ws.append((self.getId(), course))
            wb.save("ManagementSystem.xlsx")
            return True
        else:
            return False

    def loadFromExcel(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["TrainerDetails"]
        row = self.findTrainerInList(ws)
        if row:
            self.setName(ws.cell(row=row, column=2).value)
            self.setEmail(ws.cell(row=row, column=3).value)
            self.setPhoneNumber(ws.cell(row=row, column=4).value)
            wb.save("ManagementSystem.xlsx")
            return True
        else:
            return False


class Manager(Trainer):
    def __init__(self, tid="", name="", course="", email="", phone=""):
        super().__init__(tid=tid, name=name, course=course, email=email, phone=phone)