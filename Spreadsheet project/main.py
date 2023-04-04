import openpyxl
from datetime import datetime
from Trainee import Trainee
from Trainer import Trainer
from Course import Course


def traineeActions() :
    inp = input("-----TRAINEE ACTIONS-----" +
                "\n1- Add Trainee" +
                "\n2- Delete Trainee" +
                "\n3- Update Trainee" +
                "\n=> ")
    match inp :
        case "1" :  # add trainee
            trainee = Trainee(tid=input("Enter Trainee's Id: "))
            if trainee.checkExistingTrainees() :
                print("\nTrainee with that id already exists\n")
                return
            course = Course(cid=input("Enter Trainee's Course id: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id does not exist\n")
                return
            trainee.setCourse(course.getCourseId())
            trainee.setName(input("Enter trainee's name: "))
            trainee.setBackground(input("Enter trainee's background: "))
            trainee.setWorkExperience(input("Enter trainee's work experience: "))
            trainee.saveToExcel()
            print(f"\n Trainee {trainee.name} Added\n")

        case "2" :  # delete trainee
            trainee = Trainee(tid=input("Enter Trainee's Id: "))
            if not trainee.checkExistingTrainees() :
                print("\nTrainee with that id doesn't exist\n")
                return
            if trainee.deleteTrainee() :
                print("\nTrainee deleted from the list\n")
            else :
                print("\nFailed to delete Trainee\n")

        case "3" :  # update trainee
            trainee = Trainee(tid=input("Enter Trainee's Id: "))
            if not trainee.checkExistingTrainees() :
                print("\nTrainee with that id doesn't exists\n")
                return
            course = input("Enter a new course id (blank for no change): ")
            name = input("Enter new name for trainee (blank for no change): ")
            bg = input("Enter new background details for trainee (blank for no change): ")
            we = input("Enter new work experience details for trainee (blank for on change): ")
            if trainee.updateExistingTrainee(course, name, bg, we) :
                print(f"\nTrainee {trainee.traineeId} updated\n")
            else :
                print("\nFailed to update Trainee\n")

        case _ :
            print("\nInvalid input\n")


def trainerActions() :
    inp = input("-----TRAINER ACTIONS-----" +
                "\n1- Add Trainer" +
                "\n2- Delete Trainer" +
                "\n3- Update Trainer" +
                "\n4- Remove trainer from course" +
                "\n=> ")

    match inp :
        case "1" :  # add trainer
            trainer = Trainer(tid=input("Enter Trainer's Id: "))
            if trainer.checkExistingTrainer() :
                print("\nTrainer with that id already exists\n")
                return
            course = Course(cid=input("Enter Trainer's Course id: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id does not exist\n")
                return
            if course.getTrainerId():
                print("\nCourse already has a trainer assigned\n")
                return
            trainer.setCourse(course.cid)
            trainer.setName(input("Enter trainer's name: "))
            trainer.setEmail(input("Enter trainer's email: "))
            trainer.setPhoneNumber(input("Enter trainer's phone number"))
            trainer.saveToExcel()
            print(f"\n Trainer {trainer.name} Added\n")

        case "2" :  # delete trainer
            trainer = Trainer(tid=input("Enter Trainer's Id: "))
            if not trainer.checkExistingTrainer() :
                print("\nTrainer with that id doesn't exist\n")
                return
            if trainer.deleteTrainer() :
                print("\nTrainer deleted from the list\n")
            else :
                print("\nFailed to delete Trainer\n")

        case "3" :  # update trainer
            trainer = Trainer(tid=input("Enter Trainer's Id: "))
            if not trainer.checkExistingTrainer() :
                print("\nTrainer with that id doesn't exists\n")
                return
            course = input("Enter a new course id (blank for no change): ")
            name = input("Enter new name for trainer (blank for no change): ")
            email = input("Enter new email for trainer (blank for no change): ")
            phone = input("Enter new phone number details for trainer (blank for on change): ")
            if trainer.updateExistingTrainer(course, name, email, phone) :
                print(f"\nTrainer {trainer.trainerId} updated\n")
            else :
                print("\nFailed to update Trainer\n")

        case "4":  # remove trainer from course
            course = Course(cid=input("Enter course id to remove trainer from: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id doesn't exist\n")
                return
            tid = course.removeTrainer()
            if tid:
                trainer = Trainer(tid=tid)
                trainer.loadFromExcel()
                print(f"\nTrainer {trainer.name} removed from Course {course.cid}.\n")
            else:
                print(f"\nCourse {course.cid} has no assigned trainer\n")

        case _ :
            print("\nInvalid input\n")


def courseActions() :
    inp = input("-----COURSE ACTIONS-----" +
                "\n1- Add Course" +
                "\n2- Delete Course" +
                "\n3- Update course description" +
                "\n4- set Trainer for a course" +
                "\n=> ")

    match inp :
        case "1" :  # add course
            course = Course(cid=input("Enter new course's id: "))
            if course.checkExistingCourses() :
                print("\nCourse with that id already exists\n")
                return
            course.setDescription(input("Enter short description of the course: "))
            course.saveToExcel()
            print(f"\n Course {course.cid} Added\n")

        case "2" :  # delete course
            course = Course(cid=input("Enter new course's id: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id doesn't exist\n")
                return
            if course.deleteCourse() :
                print("\nCourse deleted from the list\n")
            else :
                print("\nFailed to delete Course\n")

        case "3" :  # update course
            course = Course(cid=input("Enter id of course to update: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id doesn't exist\n")
                return
            desc = input("Enter new course description (blank for no change): ")
            if course.updateCourse(desc) :
                print(f"\nCourse {course.cid} updated\n")
            else :
                print("\nFailed to update Course\n")

        case "4" :  # set trainer for course
            course = Course(cid=input("Enter id of course to set trainer for: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id doesn't exist\n")
                return
            trainer = Trainer(tid=input(f"Enter id of trainer to assign to course {course.getCourseId()}: "))
            if not trainer.checkExistingTrainer() :
                print("\nTrainer with that id doesn't exist\n")
                return
            trainer.loadFromExcel()
            if course.setTrainer(trainer.trainerId):
                print(f"\nCourse {course.getCourseId()} now has trainer {trainer.name}.\n")
            else:
                print(f"\nFailed to set trainer for course {course.getCourseId()}\n")

        case _ :
            print("\nInvalid input\n")


def saveSession(trainer, course, startTime, endTime, traineeList, attendance) :
    sheetTitle = datetime.today().strftime('%Y-%m-%d')
    wb = openpyxl.load_workbook("ManagementSystem.xlsx")
    if sheetTitle in wb.sheetnames :
        wb.remove(wb[sheetTitle])
    wb.create_sheet(sheetTitle)
    ws = wb[sheetTitle]

    ws.append(("Trainer", "Course", "Start time", "End time"))
    ws.append((trainer.name, course.cid, startTime, endTime))
    ws.insert_rows(4, amount=1)
    ws.cell(row=4, column=1).value = "Trainee Id"
    ws.cell(row=4, column=2).value = "Trainee name"
    ws.cell(row=4, column=3).value = "Attendance"
    for i, trainee in enumerate(traineeList) :
        ws.append((trainee.traineeId, trainee.name, attendance[i]))

    wb.save("ManagementSystem.xlsx")
    print(f"\nSession on {sheetTitle} saved.\n")


while True :
    inp = input("What do you want to do?" +
                "\n1- Trainee Options" +
                "\n2- Trainer Options" +
                "\n3- Course Options" +
                "\n4- Add Session" +
                "\n5- Exit" +
                "\n=> ")

    match inp :
        case "1" :  # trainee actions
            traineeActions()

        case "2" :  # trainer actions
            trainerActions()

        case "3" :  # course actions
            courseActions()

        case "4" :  # add session
            course = Course(cid=input("Enter course id to add a session for: "))
            if not course.checkExistingCourses() :
                print("\nCourse with that id doesn't exist\n")
                continue
            courseTrainerId = course.getTrainerId()
            if not courseTrainerId :
                print("\nCourse has no trainer assigned\n")
                continue
            trainer = Trainer(tid=courseTrainerId)
            if not trainer.loadFromExcel() :
                print("\nFailed to load trainer from Excel\n")
                continue

            startTime = input("Enter session start time: ")
            endTime = input("Enter session end time: ")

            traineeIdList = course.getTraineeIdList()
            traineeList = [Trainee(tid=i) for i in traineeIdList]
            for trainee in traineeList :
                trainee.loadFromExcel()
            attendance = ["P" for i in range(len(traineeList))]

            while True :
                print(f"\n---COURSE {course.getCourseId()} TRAINEES---")
                print("Id" + " " * 10 + "Name" + " "*15 + "Attendance")
                for i, trainee in enumerate(traineeList):
                    print(f"{trainee.traineeId:11} {trainee.name:18} {attendance[i]}")
                ids = input("Enter trainee IDs to mark as absent (blank for done): ").split()
                if len(ids) == 0 :
                    break
                for tid in ids:
                    for i in range(len(attendance)):
                        if traineeList[i].getId() == tid:
                            if attendance[i] == "P":
                                print(f"Trainee {traineeList[i].name} marked as absent")
                                attendance[i] = "A"
                            else:
                                print(f"Trainee {traineeList[i].name} marked as present")
                                attendance[i] = "P"

            saveSession(trainer, course, startTime, endTime, traineeList, attendance)

        case "5" :  # exit
            break
        case _ :
            print("\nInvalid input.\n")
            pass
