import openpyxl

from Course import Course


class Trainee :
    def __init__(self, tid="", name="", course="", bg="", we="") :
        self.traineeId = tid
        self.name = name
        self.course = course
        self.background = bg
        self.workExp = we

    def getId(self) :
        return self.traineeId

    def setId(self, tid) :
        self.traineeId = tid

    def getName(self) :
        return self.name

    def setName(self, name) :
        self.name = name

    def getCourse(self) :
        return self.course

    def setCourse(self, course) :
        self.course = course

    def getBackground(self) :
        return self.background

    def setBackground(self, bg) :
        self.background = bg

    def getWorkExperience(self) :
        return self.workExp

    def setWorkExperience(self, we) :
        self.workExp = we

    def findTraineeInList(self, ws) :  # returns row in the sheet that the employee is in
        values = [i for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=1).value == self.traineeId]
        if len(values) > 0 :
            return values[0]
        else :
            return False

    def checkExistingTrainees(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["ListOfTrainees"]
        if self.findTraineeInList(ws) :
            return True
        else :
            return False

    def saveToExcel(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["ListOfTrainees"]

        newRow = (self.traineeId, self.name, self.background, self.workExp)
        ws.append(newRow)

        ws = wb["MappingCourseTrainee"]
        ws.append((self.traineeId, self.course))

        wb.save("ManagementSystem.xlsx")

    def deleteTrainee(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        deleted = False

        ws = wb["ListOfTrainees"]  # remove trainee from list of trainees
        row2del = self.findTraineeInList(ws)
        if row2del:
            ws.delete_rows(row2del, 1)
            deleted = True

        ws = wb["MappingCourseTrainee"]  # remove trainee from list of course mappings
        row2del = self.findTraineeInList(ws)
        if row2del:
            ws.delete_rows(row2del, 1)
            deleted = True

        wb.save("ManagementSystem.xlsx")
        return deleted

    def updateExistingTrainee(self, course, name, bg, we) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["ListOfTrainees"]
        row = self.findTraineeInList(ws)
        if row:
            if name != "":
                ws.cell(row=row, column=2).value = name
            if bg != "":
                ws.cell(row=row, column=3).value = bg
            if we != "":
                ws.cell(row=row, column=4).value = we

            if course != "":  # TODO: Verify course exists
                c = Course(cid=course)
                if c.checkExistingCourses() :
                    ws = wb["MappingCourseTrainer"]
                    row = self.findTraineeInList(ws)
                    if row :
                        ws.cell(row=row, column=2).value = course
                    else :
                        ws.append((self.getId(), course))
            wb.save("ManagementSystem.xlsx")
            return True
        else:
            return False

    def loadFromExcel(self) :
        wb = openpyxl.load_workbook("ManagementSystem.xlsx")
        ws = wb["ListOfTrainees"]
        row = self.findTraineeInList(ws)
        if row:
            self.setName(ws.cell(row=row, column=2).value)
            self.setBackground(ws.cell(row=row, column=3).value)
            self.setWorkExperience(ws.cell(row=row, column=4).value)
            ws = wb["MappingCourseTrainee"]
            row = self.findTraineeInList(ws)
            self.setCourse(ws.cell(row=row, column=2).value)
            wb.save("ManagementSystem.xlsx")
            return True
        else:
            return False

